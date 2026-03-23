"""
export_schedule_excel.py
========================
Runs the MSE 435 scheduler and writes the results to a formatted Excel workbook.

Sheets produced:
  1. Schedule_<date>   – one sheet per day, Gantt-style room × time grid
  2. Appointments      – flat table of every assigned appointment
  3. KPIs              – key performance indicators
  4. Policy Comparison – all six policies A–F vs optimal

Usage:
    python export_schedule_excel.py --csv AppointmentDataWeek1.csv [--day 11-10-2025] [--out schedule.xlsx]
"""

import argparse
import sys
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── import scheduler internals ────────────────────────────────────────────────
sys.path.insert(0, ".")
from hospital_room_scheduler import (
    load_appointments, generate_initial_patterns, build_and_solve_master,
    column_generation, compute_kpis, ROOMS, Appointment, Pattern,
    policy_a_single_room, policy_b_cluster, policy_c_blocked_days,
    policy_d_admin_overflow, policy_e_overbook, policy_f_uncertainty,
    blocked_windows, is_available,
)

# ── colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_DARK  = "1F3864"   # navy
CLR_HEADER_MED   = "2E75B6"   # blue
CLR_HEADER_LIGHT = "D6E4F0"   # pale blue
CLR_ACCENT       = "F2C94C"   # amber (admin/blocked)
CLR_NOSHOW       = "F4B8B8"   # soft red
CLR_APPT         = "C6EFCE"   # light green (assigned)
CLR_BLOCKED_APPT = "FFE699"   # amber — appointment in a blocked window
CLR_WHITE        = "FFFFFF"
CLR_ALT_ROW      = "EBF3FB"   # alternating row tint

THIN  = Side(style="thin",   color="BFBFBF")
MED   = Side(style="medium", color="2E75B6")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _header_cell(ws, row, col, value, dark=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=True, color=CLR_WHITE, size=10)
    cell.fill      = _fill(CLR_HEADER_DARK if dark else CLR_HEADER_MED)
    cell.alignment = _center()
    cell.border    = BORDER_THIN
    return cell


# ── Sheet 1: per-day schedule ─────────────────────────────────────────────────

def _minutes_to_str(m: int) -> str:
    h, mn = divmod(m, 60)
    return f"{h:02d}:{mn:02d}"


def write_schedule_sheet(
    wb: Workbook,
    date: str,
    chosen: Dict[Tuple[str, str], "Pattern"],
    groups: Dict[Tuple[str, str], List["Appointment"]],
) -> None:
    """Write a flat schedule table for one day."""
    sheet_name = f"Schedule_{date.replace('-', '')}"
    ws = wb.create_sheet(title=sheet_name)

    # Collect all (provider, pattern) for this date
    day_entries = [
        (prov, pat)
        for (prov, d), pat in sorted(chosen.items())
        if d == date
    ]
    if not day_entries:
        ws["A1"] = f"No schedule produced for {date}"
        return

    # Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value     = f"Examination Room Schedule  –  {date}"
    title_cell.font      = _font(bold=True, color=CLR_WHITE, size=13)
    title_cell.fill      = _fill(CLR_HEADER_DARK)
    title_cell.alignment = _center()

    # Column headers
    headers = ["Provider", "Patient ID", "Appt ID", "Start", "End",
               "Duration (min)", "Room", "Appt Type", "No-Show"]
    for col, h in enumerate(headers, start=1):
        _header_cell(ws, 2, col, h, dark=False)

    col_widths = [12, 14, 9, 8, 8, 15, 7, 24, 9]
    for col, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    data_row = 3
    for prov, pat in day_entries:
        appts = sorted(
            [a for a in groups.get((prov, date), []) if a.is_active],
            key=lambda a: a.start_min,
        )
        for a in appts:
            room = pat.assignment.get(a.appt_id)
            blocks = blocked_windows(date)
            is_blocked = not is_available(a.start_min, a.end_min, blocks)

            if room is not None:
                room_str = f"ER{room}"
            elif is_blocked:
                room_str = "BLOCKED"
            else:
                room_str = "UNASSIGNED"

            row_data = [
                prov,
                a.patient_id,
                a.appt_id,
                _minutes_to_str(a.start_min),
                _minutes_to_str(a.end_min),
                a.duration,
                room_str,
                a.appt_type or "",
                "Yes" if a.no_show else "No",
            ]
            if a.no_show:
                bg = CLR_NOSHOW
            elif is_blocked:
                bg = CLR_BLOCKED_APPT
            elif data_row % 2 == 0:
                bg = CLR_ALT_ROW
            else:
                bg = CLR_WHITE

            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row, column=col, value=val)
                cell.fill      = _fill(bg)
                cell.border    = BORDER_THIN
                cell.font      = _font(size=9)
                cell.alignment = _center() if col in (1, 3, 4, 5, 6, 7, 9) else _left()
            data_row += 1

    # Summary row
    total_appts = sum(
        len([a for a in groups.get((prov, date), []) if a.is_active])
        for prov, _ in day_entries
    )
    ws.cell(row=data_row, column=1, value="TOTAL APPOINTMENTS")
    ws.cell(row=data_row, column=1).font = _font(bold=True, size=9)
    ws.cell(row=data_row, column=6, value=f'=COUNTA(F3:F{data_row-1})')
    for col in range(1, 10):
        ws.cell(row=data_row, column=col).fill   = _fill(CLR_HEADER_LIGHT)
        ws.cell(row=data_row, column=col).border = BORDER_THIN
        ws.cell(row=data_row, column=col).font   = _font(bold=True, size=9)

    ws.freeze_panes = "A3"


# ── Sheet 2: flat appointments table ─────────────────────────────────────────

def write_appointments_sheet(
    wb: Workbook,
    chosen: Dict[Tuple[str, str], "Pattern"],
    groups: Dict[Tuple[str, str], List["Appointment"]],
) -> None:
    ws = wb.create_sheet(title="Appointments")

    ws.merge_cells("A1:K1")
    ws["A1"].value     = "All Assigned Appointments"
    ws["A1"].font      = _font(bold=True, color=CLR_WHITE, size=13)
    ws["A1"].fill      = _fill(CLR_HEADER_DARK)
    ws["A1"].alignment = _center()

    headers = ["Date", "Provider", "Patient ID", "Appt ID",
               "Start", "End", "Duration (min)", "Room",
               "Appt Type", "No-Show", "Travel Cost (m)"]
    for col, h in enumerate(headers, start=1):
        _header_cell(ws, 2, col, h, dark=False)

    widths = [12, 10, 14, 9, 8, 8, 15, 7, 24, 9, 15]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 3
    for (prov, date), pat in sorted(chosen.items()):
        appts = sorted(
            [a for a in groups.get((prov, date), []) if a.is_active],
            key=lambda a: a.start_min,
        )
        for a in appts:
            room = pat.assignment.get(a.appt_id)
            blocks = blocked_windows(date)
            is_blocked = not is_available(a.start_min, a.end_min, blocks)
            if room is not None:
                room_str = f"ER{room}"
            elif is_blocked:
                room_str = "BLOCKED"
            else:
                room_str = "UNASSIGNED"
            bg = CLR_NOSHOW if a.no_show else (CLR_BLOCKED_APPT if is_blocked else (CLR_ALT_ROW if row % 2 == 0 else CLR_WHITE))
            vals = [
                date, prov, a.patient_id, a.appt_id,
                _minutes_to_str(a.start_min),
                _minutes_to_str(a.end_min),
                a.duration,
                room_str,
                a.appt_type or "",
                "Yes" if a.no_show else "No",
                round(pat.cost, 2) if a == appts[0] else "",
            ]
            for col, v in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.fill      = _fill(bg)
                cell.border    = BORDER_THIN
                cell.font      = _font(size=9)
                cell.alignment = _center() if col not in (2, 3, 9) else _left()
            row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:K{row-1}"


# ── Sheet 3: KPIs ─────────────────────────────────────────────────────────────

def write_kpi_sheet(
    wb: Workbook,
    kpis: Dict,
) -> None:
    ws = wb.create_sheet(title="KPIs")

    ws.merge_cells("A1:C1")
    ws["A1"].value     = "Key Performance Indicators – Column Generation Optimal"
    ws["A1"].font      = _font(bold=True, color=CLR_WHITE, size=13)
    ws["A1"].fill      = _fill(CLR_HEADER_DARK)
    ws["A1"].alignment = _center()

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12

    _header_cell(ws, 2, 1, "Metric",  dark=False)
    _header_cell(ws, 2, 2, "Value",   dark=False)
    _header_cell(ws, 2, 3, "Unit",    dark=False)

    kpi_rows = [
        ("Total travel distance",           "total_travel_m",                "metres"),
        ("Total room switches",              "total_room_switches",           "count"),
        ("Unassigned appointments",          "unassigned_appointments",       "count"),
        ("Provider-days solved",             "provider_days_solved",          "count"),
        ("Provider-days using single room",  "provider_days_single_room",     "count"),
        ("Avg travel per provider-day",      "avg_travel_per_provider_day",   "metres"),
    ]

    for r, (label, key, unit) in enumerate(kpi_rows, start=3):
        bg = CLR_ALT_ROW if r % 2 == 0 else CLR_WHITE
        val = kpis.get(key, 0)
        val_str = f"{val:.2f}" if isinstance(val, float) else val

        for col, v in enumerate([label, val_str, unit], start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill      = _fill(bg)
            cell.border    = BORDER_THIN
            cell.font      = _font(size=10)
            cell.alignment = _left() if col == 1 else _center()


# ── Sheet 4: Policy comparison ───────────────────────────────────────────────

def write_policy_sheet(
    wb: Workbook,
    groups: Dict[Tuple[str, str], List["Appointment"]],
    optimal_kpis: Dict,
) -> None:
    ws = wb.create_sheet(title="Policy Comparison")

    ws.merge_cells("A1:G1")
    ws["A1"].value     = "Policy Comparison  –  Policies A–F  vs  Column Generation Optimal"
    ws["A1"].font      = _font(bold=True, color=CLR_WHITE, size=13)
    ws["A1"].fill      = _fill(CLR_HEADER_DARK)
    ws["A1"].alignment = _center()

    metrics = [
        ("Total Travel (m)",          "total_travel_m"),
        ("Room Switches",              "total_room_switches"),
        ("Avg Travel/Provider-Day (m)","avg_travel_per_provider_day"),
        ("Provider-Days Solved",       "provider_days_solved"),
        ("Single-Room Days",           "provider_days_single_room"),
    ]
    col_headers = ["Policy"] + [m[0] for m in metrics]
    for col, h in enumerate(col_headers, start=1):
        _header_cell(ws, 2, col, h, dark=False)

    ws.column_dimensions["A"].width = 28
    for col in range(2, len(col_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    policies = [
        ("Optimal (CG)",          None),
        ("A – Single Room (day)", lambda g: policy_a_single_room(g, week_lock=False)),
        ("A – Single Room (week)",lambda g: policy_a_single_room(g, week_lock=True)),
        ("B – Cluster (≤3 m)",    lambda g: policy_b_cluster(g, proximity_threshold=3.0)),
        ("B – Cluster (≤5 m)",    lambda g: policy_b_cluster(g, proximity_threshold=5.0)),
        ("C – Blocked Days",      lambda g: policy_c_blocked_days(g)),
        ("D – Admin Overflow ON", lambda g: policy_d_admin_overflow(g, allow_admin_overflow=True)),
        ("D – Admin Overflow OFF",lambda g: policy_d_admin_overflow(g, allow_admin_overflow=False)),
        ("E – Overbook (10% NS)", lambda g: policy_e_overbook(g)),
        ("F – Uncertainty σ=20%", lambda g: policy_f_uncertainty(g)),
    ]

    for r, (label, fn) in enumerate(policies, start=3):
        if fn is None:
            kpi = optimal_kpis
            bg  = CLR_HEADER_LIGHT
            bold = True
        else:
            _, chosen = fn(groups)
            kpi  = compute_kpis(chosen, groups)
            bg   = CLR_ALT_ROW if r % 2 == 0 else CLR_WHITE
            bold = False

        row_vals = [label] + [
            round(kpi.get(key, 0), 2) if isinstance(kpi.get(key, 0), float)
            else kpi.get(key, 0)
            for _, key in metrics
        ]
        for col, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.fill      = _fill(bg)
            cell.border    = BORDER_THIN
            cell.font      = _font(bold=bold, size=10)
            cell.alignment = _left() if col == 1 else _center()

    ws.freeze_panes = "B3"


# ── Main ──────────────────────────────────────────────────────────────────────

def main(csv_path: str, day_filter: Optional[str], out_path: str,
         cg_iters: int = 1) -> None:

    print(f"Loading appointments from {csv_path} ...")
    all_appts = load_appointments(csv_path)

    groups = defaultdict(list)
    for a in all_appts:
        groups[(a.provider, a.date)].append(a)

    if day_filter:
        groups = {k: v for k, v in groups.items() if k[1] == day_filter}
        print(f"  Filtered to {day_filter} — {len(groups)} provider-day groups")

    # Generate initial patterns
    print("Generating initial patterns ...")
    patterns_map = {}
    for (prov, date), appts in groups.items():
        active = [a for a in appts if a.is_active]
        if active:
            patterns_map[(prov, date)] = generate_initial_patterns(active, prov, date)

    # Column generation
    if cg_iters > 0:
        print(f"Running column generation ({cg_iters} iterations) ...")
        patterns_map = column_generation(groups, patterns_map, cg_iters)

    # Solve MIP
    print("Solving integer master problem ...")
    obj, chosen = build_and_solve_master(groups, patterns_map, integer=True)
    print(f"  Optimal total travel: {obj:.2f} m")

    kpis = compute_kpis(chosen, groups)

    # Build workbook
    print("Writing Excel workbook ...")
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    # One schedule sheet per date
    dates = sorted({date for (_, date) in chosen})
    for date in dates:
        write_schedule_sheet(wb, date, chosen, groups)

    write_appointments_sheet(wb, chosen, groups)
    write_kpi_sheet(wb, kpis)
    write_policy_sheet(wb, groups, kpis)

    wb.save(out_path)
    print(f"  Saved → {out_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export schedule to Excel")
    parser.add_argument("--csv",  default="AppointmentDataWeek1.csv")
    parser.add_argument("--day",  default=None)
    parser.add_argument("--out",  default="schedule_output2.xlsx")
    parser.add_argument("--cg-iters", type=int, default=1)
    args = parser.parse_args()
    main(args.csv, args.day, args.out, args.cg_iters)
    
    # The use of GEN-AI was used to assist exporting the output of the model in a well-formatted Excel workbook, including the design of the sheets, the formatting, and the policy comparison.